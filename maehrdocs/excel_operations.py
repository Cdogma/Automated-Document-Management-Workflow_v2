"""
Excel-Operationen für MaehrDocs
Enthält Funktionen zum Schreiben, Suchen und Aktualisieren von Excel-Daten.
"""

import os
import logging
from datetime import datetime
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink

def write_document_row(sheet, row, doc_info, file_path, columns):
    """
    Schreibt eine Zeile mit Dokumentinformationen in das Arbeitsblatt.
    
    Args:
        sheet: Das openpyxl-Arbeitsblatt
        row (int): Zeilennummer
        doc_info (dict): Extrahierte Dokumenteninformationen
        file_path (str): Pfad zur PDF-Datei (für den Hyperlink)
        columns (list): Liste der Spaltenüberschriften
    """
    # Mapping von JSON-Schlüsseln zu Spalten
    field_mapping = {
        "datum": "Datum",
        "dokumenttyp": "Typ",
        "absender": "Absender",
        "betreff": "Betreff",
        "netto": "Netto",
        "mwst_betrag": "MwSt",
        "brutto": "Brutto",
        "waehrung": "Währung",
        "rechnungsnummer": "Rechnungsnummer",
        "art": "Art",
        "bezahlt": "Bezahlt",
        "steuersatz": "Steuersatz"
    }
    
    # Werte eintragen
    for json_key, excel_col in field_mapping.items():
        value = doc_info.get(json_key, "")
        col_idx = columns.index(excel_col) + 1
        
        # Für numerische Felder sicherstellen, dass sie als Zahlen gespeichert werden
        if excel_col in ["Netto", "MwSt", "Brutto"]:
            try:
                value = float(value) if value else 0.0
            except (ValueError, TypeError):
                value = 0.0
        
        sheet.cell(row=row, column=col_idx, value=value)
    
    # Datei-Link setzen
    set_hyperlink(sheet, row, file_path, columns)
    
    # Lexware-Status (standardmäßig "pending")
    lexware_col = columns.index("Lexware-Status") + 1
    sheet.cell(row=row, column=lexware_col, value="pending")

def set_hyperlink(sheet, row, file_path, columns):
    """
    Erstellt einen Hyperlink zur PDF-Datei.
    
    Args:
        sheet: Das openpyxl-Arbeitsblatt
        row (int): Zeilennummer
        file_path (str): Pfad zur PDF-Datei
        columns (list): Liste der Spaltenüberschriften
    """
    link_col = columns.index("Datei-Link") + 1
    cell = sheet.cell(row=row, column=link_col)
    
    # Hyperlink erstellen
    if file_path:
        # Relationen-ID für Hyperlink
        rel_id = f"rId{row}"
        
        # Absolute Pfadangabe
        absolute_path = os.path.abspath(file_path)
        
        # Hyperlink erstellen und in Zelle setzen
        hyperlink = Hyperlink(display="📎", ref=f"{cell.coordinate}", id=rel_id, target=absolute_path, location=None)
        sheet.hyperlinks.append(hyperlink)
        
        # Zelleninhalt setzen
        cell.value = "📎"
        cell.style = "Hyperlink"

def search_documents(excel_path, criteria, columns, logger=None):
    """
    Sucht nach Dokumenten in der Excel-Tabelle anhand von Kriterien.
    
    Args:
        excel_path (str): Pfad zur Excel-Datei
        criteria (dict): Suchkriterien, z.B. {"Absender": "Telekom", "Bezahlt": "nein"}
        columns (list): Liste der Spaltenüberschriften
        logger (Logger, optional): Logger für eventuelle Fehlermeldungen
        
    Returns:
        list: Liste der gefundenen Zeilen als Dictionaries
    """
    try:
        # Prüfen, ob die Datei existiert
        if not os.path.exists(excel_path):
            if logger:
                logger.warning(f"Excel-Datei nicht gefunden: {excel_path}")
            return []
        
        # Excel-Datei öffnen
        workbook = openpyxl.load_workbook(excel_path, read_only=True)
        
        # Arbeitsblatt auswählen
        sheet_name = "Alle_Rechnungen"
        if sheet_name not in workbook.sheetnames:
            if logger:
                logger.warning(f"Arbeitsblatt nicht gefunden: {sheet_name}")
            return []
        
        sheet = workbook[sheet_name]
        
        # Ergebnisliste
        results = []
        
        # Spaltenindizes für die Kriterien bestimmen
        criteria_indices = {}
        for key in criteria:
            if key in columns:
                criteria_indices[key] = columns.index(key) + 1
        
        # Zeilen durchsuchen
        for row_idx, row in enumerate(sheet.rows, start=1):
            # Überschriftenzeile überspringen
            if row_idx == 1:
                continue
            
            # Prüfen, ob die Zeile allen Kriterien entspricht
            match = True
            for key, col_idx in criteria_indices.items():
                cell_value = str(row[col_idx - 1].value).lower() if row[col_idx - 1].value is not None else ""
                criteria_value = str(criteria[key]).lower()
                
                if criteria_value not in cell_value:
                    match = False
                    break
            
            # Wenn alle Kriterien erfüllt sind, Zeile zur Ergebnisliste hinzufügen
            if match:
                row_data = {}
                for col_idx, col_name in enumerate(columns):
                    row_data[col_name] = row[col_idx].value
                
                results.append(row_data)
        
        return results
        
    except Exception as e:
        if logger:
            logger.error(f"Fehler bei der Suche nach Dokumenten: {str(e)}")
        return []

def update_payment_status(excel_path, row_idx, payment_status, columns, logger=None):
    """
    Aktualisiert den Bezahltstatus eines Dokuments.
    
    Args:
        excel_path (str): Pfad zur Excel-Datei
        row_idx (int): Index der zu aktualisierenden Zeile
        payment_status (str): Neuer Bezahltstatus ("ja", "nein", "unbekannt")
        columns (list): Liste der Spaltenüberschriften
        logger (Logger, optional): Logger für eventuelle Fehlermeldungen
        
    Returns:
        bool: True bei Erfolg, False bei Fehler
    """
    try:
        # Prüfen, ob die Datei existiert
        if not os.path.exists(excel_path):
            if logger:
                logger.warning(f"Excel-Datei nicht gefunden: {excel_path}")
            return False
        
        # Excel-Datei öffnen
        workbook = openpyxl.load_workbook(excel_path)
        
        # Arbeitsblatt auswählen
        sheet_name = "Alle_Rechnungen"
        if sheet_name not in workbook.sheetnames:
            if logger:
                logger.warning(f"Arbeitsblatt nicht gefunden: {sheet_name}")
            return False
        
        sheet = workbook[sheet_name]
        
        # Prüfen, ob die Zeile existiert
        if row_idx < 2 or row_idx > sheet.max_row:
            if logger:
                logger.warning(f"Ungültiger Zeilenindex: {row_idx}")
            return False
        
        # Bezahltspalte aktualisieren
        bezahlt_col = columns.index("Bezahlt") + 1
        sheet.cell(row=row_idx, column=bezahlt_col, value=payment_status)
        
        # Excel-Datei speichern
        workbook.save(excel_path)
        
        if logger:
            logger.info(f"Bezahltstatus aktualisiert: Zeile {row_idx}, Status {payment_status}")
        return True
        
    except Exception as e:
        if logger:
            logger.error(f"Fehler beim Aktualisieren des Bezahltstatus: {str(e)}")
        return False

def export_excel_to_pdf(excel_path, output_path, logger=None):
    """
    Exportiert die Excel-Tabelle als PDF.
    
    Diese Funktion ist nur ein Platzhalter, da der direkte Export als PDF
    mit openpyxl nicht möglich ist. In einer realen Implementierung könnte
    man Excel über COM-Automatisierung oder andere Bibliotheken verwenden.
    
    Args:
        excel_path (str): Pfad zur Excel-Datei
        output_path (str): Pfad zur Ausgabe-PDF
        logger (Logger, optional): Logger für eventuelle Warnungen
        
    Returns:
        bool: True bei Erfolg, False bei Fehler
    """
    if logger:
        logger.warning("PDF-Export wird aktuell nicht unterstützt.")
        logger.info("Tipp: Sie können die Excel-Datei manuell als PDF exportieren.")
    return False