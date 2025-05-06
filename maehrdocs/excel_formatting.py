"""
Formatierungsfunktionen für Excel-Dateien in MaehrDocs
Enthält Funktionen zum Einrichten und Formatieren von Excel-Arbeitsblättern.
"""

import logging
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def setup_sheet(sheet, columns):
    """
    Richtet ein Arbeitsblatt mit Spaltenüberschriften und Formatierung ein.
    
    Args:
        sheet: Das einzurichtende openpyxl-Arbeitsblatt
        columns (list): Liste der Spaltenüberschriften
    """
    # Spaltenüberschriften eintragen
    for col, header in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Spaltenbreiten anpassen
    set_column_widths(sheet, columns)
    
    # Tabelle erstellen
    create_table(sheet, columns)

def set_column_widths(sheet, columns):
    """
    Setzt die Breiten der Spalten basierend auf deren Inhalt.
    
    Args:
        sheet: Das zu formatierende openpyxl-Arbeitsblatt
        columns (list): Liste der Spaltenüberschriften
    """
    column_widths = {
        "Datum": 12,
        "Typ": 15,
        "Absender": 30,
        "Betreff": 40,
        "Netto": 12,
        "MwSt": 12,
        "Brutto": 12,
        "Währung": 8,
        "Rechnungsnummer": 20,
        "Art": 10,
        "Bezahlt": 10,
        "Steuersatz": 10,
        "Datei-Link": 10,
        "Lexware-Status": 15,
        "Kommentar": 30
    }
    
    for col, header in enumerate(columns, start=1):
        width = column_widths.get(header, 15)
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = width

def create_table(sheet, columns):
    """
    Erstellt eine formatierte Tabelle im Arbeitsblatt.
    
    Args:
        sheet: Das openpyxl-Arbeitsblatt
        columns (list): Liste der Spaltenüberschriften
    """
    # Tabelle erstellen
    table = Table(displayName="Rechnungen", ref=f"A1:{get_column_letter(len(columns))}1")
    
    # Tabellenstil festlegen
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    
    # Tabelle zum Blatt hinzufügen
    sheet.add_table(table)

def update_table(sheet, columns, logger=None):
    """
    Aktualisiert die Tabellenreferenz, um neue Zeilen einzuschließen.
    
    Args:
        sheet: Das zu aktualisierende openpyxl-Arbeitsblatt
        columns (list): Liste der Spaltenüberschriften
        logger (Logger, optional): Logger für eventuelle Warnungen
    """
    try:
        # Suche nach der Tabelle
        for table in sheet.tables.values():
            # Aktualisiere den Tabellenbereich
            table.ref = f"A1:{get_column_letter(len(columns))}{sheet.max_row}"
    except Exception as e:
        if logger:
            logger.warning(f"Tabelle konnte nicht aktualisiert werden: {str(e)}")
        else:
            print(f"Warnung: Tabelle konnte nicht aktualisiert werden: {str(e)}")