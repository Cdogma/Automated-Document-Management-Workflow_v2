"""
Excel-Integration für MaehrDocs
Enthält die Funktionalität zum Hinzufügen und Aktualisieren von Dokumenteninformationen
in Excel-Tabellen für die Dokumentenverwaltung und -analyse.
"""

import os
import logging
import datetime
from pathlib import Path

class ExcelWriter:
    """
    Klasse zum Schreiben von Dokumenteninformationen in Excel-Tabellen.
    
    Diese Klasse ist verantwortlich für:
    - Erstellen oder Öffnen von Excel-Tabellen für Dokumente
    - Hinzufügen neuer Dokumente zur Tabelle
    - Aktualisieren von Dokumentinformationen
    - Einrichten von Hyperlinks zu den Dokumenten
    - Automatisches Anlegen erforderlicher Ordnerstrukturen
    """
    
    def __init__(self, config):
        """
        Initialisiert den ExcelWriter mit der Konfiguration.
        
        Args:
            config (dict): Konfigurationsdaten für die Excel-Integration
        """
        self.config = config
        self.logger = logging.getLogger(__name__)
        
        # Versuche erforderliche Bibliotheken zu importieren
        try:
            import openpyxl
            self.openpyxl = openpyxl
            self.is_available = True
        except ImportError:
            self.logger.warning("openpyxl nicht installiert. Excel-Integration deaktiviert.")
            self.is_available = False
    
    def add_document(self, doc_info, file_path):
        """
        Fügt ein Dokument zur Excel-Tabelle hinzu.
        
        Diese Methode öffnet die entsprechende Excel-Datei für das Jahr des Dokuments
        und fügt eine neue Zeile mit den Dokumentinformationen hinzu.
        
        Args:
            doc_info (dict): Extrahierte Dokumenteninformationen
            file_path (str): Pfad zur verarbeiteten PDF-Datei (für den Hyperlink)
            
        Returns:
            bool: True bei Erfolg, False bei Fehler
        """
        if not self.is_available:
            return False
            
        try:
            # Datum des Dokuments extrahieren
            date_str = doc_info.get('datum', '')
            if not date_str:
                self.logger.warning("Datum des Dokuments fehlt. Excel-Eintrag übersprungen.")
                return False
                
            # Datum parsen und Jahr extrahieren
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
            year = date.strftime("%Y")
            
            # Excel-Datei für das Jahr bestimmen
            excel_dir = self.config.get('paths', {}).get('excel_dir', '')
            if not excel_dir:
                # Falls excel_dir nicht konfiguriert ist, nehmen wir den output_dir
                excel_dir = self.config.get('paths', {}).get('output_dir', '')
            
            # Prüfen, ob das Verzeichnis existiert
            if not os.path.exists(excel_dir):
                os.makedirs(excel_dir)
            
            # Excel-Datei für das Jahr
            excel_file = os.path.join(excel_dir, f"Rechnungen_{year}.xlsx")
            
            # Prüfen, ob die Datei bereits existiert
            if not os.path.exists(excel_file):
                self._create_new_excel_file(excel_file)
            
            # Excel-Datei öffnen
            workbook = self.openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            # Neue Zeile hinzufügen
            row_data = self._prepare_row_data(doc_info, file_path)
            self._add_row_to_excel(sheet, row_data)
            
            # Datei speichern
            workbook.save(excel_file)
            
            self.logger.info(f"Dokument zur Excel-Tabelle {excel_file} hinzugefügt.")
            return True
            
        except Exception as e:
            self.logger.error(f"Fehler beim Hinzufügen des Dokuments zur Excel-Tabelle: {str(e)}")
            return False
    
    def _create_new_excel_file(self, excel_file):
        """
        Erstellt eine neue Excel-Datei mit den erforderlichen Spalten.
        
        Args:
            excel_file (str): Pfad zur zu erstellenden Excel-Datei
        
        Returns:
            bool: True bei Erfolg, False bei Fehler
        """
        try:
            # Neue Arbeitsmappe erstellen
            workbook = self.openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Rechnungen"
            
            # Spaltenüberschriften definieren
            headers = [
                "Datum", "Typ", "Absender", "Betreff", "Netto", "MwSt", "Brutto", 
                "Währung", "Rechnungsnummer", "Art", "Bezahlt", "Steuersatz", 
                "Link", "Lexware-Status"
            ]
            
            # Spaltenüberschriften in die erste Zeile einfügen
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                # Kopfzeile formatieren (Fett, zentriert)
                cell.font = self.openpyxl.styles.Font(bold=True)
                cell.alignment = self.openpyxl.styles.Alignment(horizontal='center')
            
            # Autofilter aktivieren
            sheet.auto_filter.ref = f"A1:{chr(65 + len(headers) - 1)}1"
            
            # Datei speichern
            workbook.save(excel_file)
            
            self.logger.info(f"Neue Excel-Datei erstellt: {excel_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"Fehler beim Erstellen der Excel-Datei: {str(e)}")
            return False
    
    def _prepare_row_data(self, doc_info, file_path):
        """
        Bereitet die Daten für eine neue Zeile in der Excel-Tabelle vor.
        
        Args:
            doc_info (dict): Extrahierte Dokumenteninformationen
            file_path (str): Pfad zur verarbeiteten PDF-Datei (für den Hyperlink)
            
        Returns:
            dict: Vorbereitete Zeilendaten
        """
        # Extrahiere die benötigten Werte aus doc_info mit Standardwerten
        date_str = doc_info.get('datum', '')
        doc_type = doc_info.get('dokumenttyp', 'Dokument')
        sender = doc_info.get('absender', '')
        subject = doc_info.get('betreff', '')
        
        # Finanzinformationen extrahieren (falls verfügbar)
        kennzahlen = doc_info.get('kennzahlen', {})
        brutto = kennzahlen.get('brutto', '')
        netto = kennzahlen.get('netto', '')
        mwst = kennzahlen.get('mwst_betrag', '')
        currency = kennzahlen.get('währung', '€')
        invoice_number = kennzahlen.get('rechnungsnummer', '')
        tax_rate = kennzahlen.get('steuersatz', '')
        
        # Art des Dokuments (business, privat, bar)
        art = doc_info.get('art', 'privat')
        
        # Zahlung und Status
        paid = "Nein"  # Standardwert
        lexware_status = "pending"  # Standardwert
        
        # Datei-Link
        file_name = os.path.basename(file_path)
        file_link = file_path  # Vollständiger Pfad für den Hyperlink
        
        # Zeilendaten als Dictionary zurückgeben
        return {
            "datum": date_str,
            "typ": doc_type,
            "absender": sender,
            "betreff": subject,
            "netto": netto,
            "mwst": mwst,
            "brutto": brutto,
            "währung": currency,
            "rechnungsnummer": invoice_number,
            "art": art,
            "bezahlt": paid,
            "steuersatz": tax_rate,
            "lexware_status": lexware_status,
            "link": file_link
        }
    
    def _add_row_to_excel(self, sheet, row_data):
        """
        Fügt eine neue Zeile zur Excel-Tabelle hinzu.
        
        Args:
            sheet: Das aktive Arbeitsblatt
            row_data (dict): Die einzufügenden Zeilendaten
            
        Returns:
            bool: True bei Erfolg, False bei Fehler
        """
        try:
            # Nächste freie Zeile bestimmen
            next_row = sheet.max_row + 1
            
            # Spaltenreihenfolge (muss mit den Spaltenüberschriften übereinstimmen)
            columns = [
                "datum", "typ", "absender", "betreff", "netto", "mwst", "brutto", 
                "währung", "rechnungsnummer", "art", "bezahlt", "steuersatz", 
                "link", "lexware_status"
            ]
            
            # Daten in die Zeile einfügen
            for col_idx, col_name in enumerate(columns, 1):
                cell = sheet.cell(row=next_row, column=col_idx)
                value = row_data.get(col_name, '')
                
                # Für den Link eine Excel-Hyperlink-Formel erstellen
                if col_name == "link" and value:
                    file_path = value
                    file_name = os.path.basename(file_path)
                    cell.hyperlink = file_path
                    cell.value = file_name
                    cell.style = "Hyperlink"
                else:
                    cell.value = value
            
            return True
            
        except Exception as e:
            self.logger.error(f"Fehler beim Hinzufügen der Zeile zur Excel-Tabelle: {str(e)}")
            return False